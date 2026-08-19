#!/usr/bin/env python3
"""
merge_room_timetable.py
=======================

Merge a Room-wise timetable so each physical room is one row.

The room-wise export splits a room across its associative sections, so one
room shows up several times:

    C009-MA, C009-A, C009-B, C009-C   ->   C009

This is the standalone equivalent of the Room Merger at /converter/rooms.
It reads the export directly, so it needs no database and no running app.

Input
-----
The room-wise file (.csv or .xlsx), in its usual wide shape:

    Room No | mon1 | mon2 | ... | mon11 | tue1 | ... | sat11
    C009-A  | ...  |  -   | ... |  ...  |  -   | ... |  ...

A free slot is "-" or blank. Column names are matched case- and
space-insensitively, so "Room No", "ROOMNO" and "room_no" all work, as do
"mon1", "Mon 1" and "MON-1".

Output
------
An .xlsx workbook with four sheets:

    Merged Timetable   one row per physical room, mon1..sat11
    Merge Map          which sub-rooms merged into which room
    Multi-class Cells  cells where sub-rooms held different classes
    Summary            the counts printed at the end of the run

Merging is not always lossless. Most duplicated cells hold the identical
class, but some sub-rooms carry different sections of one course - one room
may hold SEC 66, 67, 71, 72 and 74 of the same course in the same period.
Every distinct label is kept and joined with --sep, and those cells are
listed on their own sheet. Nothing is dropped silently.

Usage
-----
    pip install openpyxl

    python scripts/merge_room_timetable.py room_timetable.csv
    python scripts/merge_room_timetable.py room_timetable.xlsx -o merged.xlsx
    python scripts/merge_room_timetable.py in.csv --max-hour 11 --sep " / "
    python scripts/merge_room_timetable.py in.csv --preview 15
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat"]

# Associative-section suffix: C009-MA, C009-A ... C009-F, and the odd -AB/-CD.
SECTION_SUFFIX = re.compile(r"-(MA|AB|CD|[A-F])$", re.IGNORECASE)

# A descriptive tail such as "F102-CHEMISTRY LAB" is the same room as F102.
# Only a tail of 3+ characters counts, so "-A" and "-MA" stay section suffixes.
DESCRIPTIVE = re.compile(r"^([A-Z]{1,3}\s?\d{2,4}[A-Z]?\d?)\s*-\s*(.{3,})$")

FREE_MARKERS = {"", "-", "--", "nil", "none"}


# --------------------------------------------------------------------------
# room names
# --------------------------------------------------------------------------
def canonical_room(raw: str) -> str:
    """Upper-case, trimmed, with a descriptive tail reduced to the room code."""
    s = str(raw or "").strip().upper()
    m = DESCRIPTIVE.match(s)
    if not m:
        return s
    tail = m.group(2).strip()
    if re.fullmatch(r"MA|AB|CD|[A-F]", tail):
        return s
    return re.sub(r"\s+", "", m.group(1))


def base_room(raw: str) -> str:
    """The physical room: strip every associative-section suffix."""
    s = canonical_room(raw)
    while SECTION_SUFFIX.search(s):
        s = SECTION_SUFFIX.sub("", s)
    return s


# --------------------------------------------------------------------------
# reading the source
# --------------------------------------------------------------------------
def _norm_key(name: str) -> str:
    """Fold a header to a comparable key: 'Mon 1' and 'MON-1' both -> 'mon1'."""
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


def read_rows(path: Path) -> list[dict]:
    """Rows as dicts, from .csv or .xlsx."""
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Reading .xlsx needs openpyxl:  pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = ["" if c is None else str(c) for c in next(rows_iter)]
        except StopIteration:
            return []
        out = []
        for cells in rows_iter:
            if cells is None or all(c is None or str(c).strip() == "" for c in cells):
                continue
            out.append({
                header[i]: ("" if cells[i] is None else str(cells[i]))
                for i in range(min(len(header), len(cells)))
            })
        wb.close()
        return out

    with path.open(newline="", encoding="utf-8-sig") as fh:
        return [dict(r) for r in csv.DictReader(fh)]


def find_room_column(headers: list[str]) -> str:
    """The column holding the room name."""
    wanted = ("roomno", "room", "roomnumber", "roomname")
    for h in headers:
        if _norm_key(h) in wanted:
            return h
    for h in headers:                       # fall back to anything room-ish
        if "room" in _norm_key(h):
            return h
    sys.exit(
        "Could not find a room column. Expected something like 'Room No'.\n"
        f"Columns seen: {', '.join(headers)}"
    )


def find_slot_columns(headers: list[str], max_hour: int) -> list[tuple[str, str, int]]:
    """[(header, day, hour)] for every day+period column within max_hour."""
    slots = []
    for h in headers:
        key = _norm_key(h)
        m = re.fullmatch(r"(mon|tue|wed|thu|fri|sat|sun)(\d{1,2})", key)
        if not m:
            continue
        day, hour = m.group(1), int(m.group(2))
        if day == "sun" or not 1 <= hour <= max_hour:
            continue
        slots.append((h, day, hour))
    slots.sort(key=lambda s: (DAYS.index(s[1]), s[2]))
    return slots


# --------------------------------------------------------------------------
# merging
# --------------------------------------------------------------------------
def merge(rows: list[dict], room_col: str, slots: list[tuple[str, str, int]]) -> dict:
    """Collapse sub-rooms into physical rooms, keeping every distinct label."""
    # room -> {"variants": set, "cells": {(day, hour): {"labels": [...], "rows": n}}}
    merged: dict[str, dict] = {}
    source_names: set[str] = set()
    used_rows = 0

    for row in rows:
        raw = str(row.get(room_col, "") or "").strip()
        if not raw:
            continue
        source_names.add(raw)
        base = base_room(raw)
        if not base:
            continue

        entry = merged.setdefault(base, {"variants": set(), "cells": {}})
        entry["variants"].add(raw)

        for header, day, hour in slots:
            value = str(row.get(header, "") or "").strip()
            if value.lower() in FREE_MARKERS:
                continue
            used_rows += 1
            cell = entry["cells"].setdefault((day, hour), {"labels": [], "rows": 0})
            cell["rows"] += 1
            if value not in cell["labels"]:
                cell["labels"].append(value)

    for entry in merged.values():
        entry["variants"] = sorted(entry["variants"])
        for cell in entry["cells"].values():
            cell["labels"].sort()

    return {"rooms": merged, "source_names": source_names, "used_cells": used_rows}


def room_sort_key(name: str):
    """Sort C9 before C10, and keep 'A 301' next to 'A306' despite the space."""
    tidy = re.sub(r"\s+", "", name)
    return [int(p) if p.isdigit() else p for p in re.split(r"(\d+)", tidy)]


# --------------------------------------------------------------------------
# writing
# --------------------------------------------------------------------------
def write_workbook(out_path: Path, result: dict, slots, sep: str, stats: dict) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.exit("Writing .xlsx needs openpyxl:  pip install openpyxl")

    rooms = result["rooms"]
    order = sorted(rooms, key=room_sort_key)
    col_keys = [(d, h) for _, d, h in slots]

    wb = Workbook()
    bold = Font(bold=True)
    amber = PatternFill("solid", fgColor="FDE8C8")

    ws = wb.active
    ws.title = "Merged Timetable"
    ws.append(["Room No"] + [f"{d}{h}" for d, h in col_keys])
    for c in ws[1]:
        c.font = bold
    for name in order:
        cells = rooms[name]["cells"]
        line = [name]
        for key in col_keys:
            cell = cells.get(key)
            line.append(sep.join(cell["labels"]) if cell else "-")
        ws.append(line)
        for i, key in enumerate(col_keys, start=2):
            cell = cells.get(key)
            if cell and len(cell["labels"]) > 1:
                ws.cell(row=ws.max_row, column=i).fill = amber
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18

    ws2 = wb.create_sheet("Merge Map")
    ws2.append(["Merged Room", "Merged From", "Sub-rooms",
                "Busy cells", "Cells holding more than one class"])
    for c in ws2[1]:
        c.font = bold
    for name in order:
        entry = rooms[name]
        multi = sum(1 for c in entry["cells"].values() if len(c["labels"]) > 1)
        ws2.append([name, " | ".join(entry["variants"]), len(entry["variants"]),
                    len(entry["cells"]), multi])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 60

    ws3 = wb.create_sheet("Multi-class Cells")
    ws3.append(["Room", "Day", "Period", "Classes in this cell", "Labels"])
    for c in ws3[1]:
        c.font = bold
    any_multi = False
    for name in order:
        for (day, hour), cell in sorted(rooms[name]["cells"].items(),
                                        key=lambda kv: (DAYS.index(kv[0][0]), kv[0][1])):
            if len(cell["labels"]) > 1:
                any_multi = True
                ws3.append([name, day, hour, len(cell["labels"]),
                            " | ".join(cell["labels"])])
    if not any_multi:
        ws3.append(["No cell held more than one class"])
    ws3.column_dimensions["E"].width = 80

    ws4 = wb.create_sheet("Summary")
    ws4.append(["Metric", "Value"])
    for c in ws4[1]:
        c.font = bold
    for k, v in stats.items():
        ws4.append([k, v])
    ws4.column_dimensions["A"].width = 42

    wb.save(out_path)


# --------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(
        description="Merge a Room-wise timetable so each physical room is one row.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Example:\n  python scripts/merge_room_timetable.py room_timetable.csv",
    )
    ap.add_argument("input", type=Path, help="Room-wise timetable (.csv or .xlsx)")
    ap.add_argument("-o", "--output", type=Path,
                    help="output .xlsx (default: <input>-merged.xlsx)")
    ap.add_argument("--max-hour", type=int, default=11,
                    help="keep periods 1..N, drop the rest (default 11)")
    ap.add_argument("--sep", default=" | ",
                    help="joins several classes in one cell (default ' | ')")
    ap.add_argument("--preview", type=int, default=10,
                    help="rooms to print as a preview, 0 for none (default 10)")
    ap.add_argument("--drop-empty", action="store_true",
                    help="omit rooms with no class in the periods kept")
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"No such file: {args.input}")

    rows = read_rows(args.input)
    if not rows:
        sys.exit("The file has no data rows.")

    headers = list(rows[0].keys())
    room_col = find_room_column(headers)
    slots = find_slot_columns(headers, args.max_hour)
    if not slots:
        sys.exit(
            "No day/period columns found. Expected headers like 'mon1' … 'sat11'.\n"
            f"Columns seen: {', '.join(headers)}"
        )

    dropped_cols = [
        h for h in headers
        if re.fullmatch(r"(mon|tue|wed|thu|fri|sat|sun)(\d{1,2})", _norm_key(h))
        and h not in {s[0] for s in slots}
    ]

    result = merge(rows, room_col, slots)
    if args.drop_empty:
        result["rooms"] = {k: v for k, v in result["rooms"].items() if v["cells"]}
    rooms = result["rooms"]

    busy_cells = sum(len(r["cells"]) for r in rooms.values())
    multi_cells = sum(1 for r in rooms.values()
                      for c in r["cells"].values() if len(c["labels"]) > 1)
    rooms_multi = sum(1 for r in rooms.values()
                      if any(len(c["labels"]) > 1 for c in r["cells"].values()))
    unchanged = sum(1 for r in rooms.values() if len(r["variants"]) == 1)
    empty = sum(1 for r in rooms.values() if not r["cells"])

    stats = {
        "Source file": args.input.name,
        "Source data rows": len(rows),
        "Room names in source": len(result["source_names"]),
        "Merged rooms": len(rooms),
        "Rooms that had no sub-rooms": unchanged,
        "Periods kept": f"1-{args.max_hour}",
        "Period columns dropped": len(dropped_cols),
        "Filled cells read": result["used_cells"],
        "Busy cells after merge": busy_cells,
        "Rooms with nothing in periods kept": empty,
        "Duplicate cells collapsed": result["used_cells"] - busy_cells,
        "Cells holding more than one class": multi_cells,
        "Rooms affected by those": rooms_multi,
        "Separator": args.sep,
    }

    out = args.output or args.input.with_name(args.input.stem + "-merged.xlsx")
    write_workbook(out, result, slots, args.sep, stats)

    width = max(len(k) for k in stats)
    print()
    for k, v in stats.items():
        print(f"  {k.ljust(width)} : {v}")

    if args.preview:
        print(f"\n  Preview - first {args.preview} merged rooms:")
        for name in sorted(rooms, key=room_sort_key)[: args.preview]:
            entry = rooms[name]
            note = (f"{len(entry['variants'])} sub-rooms"
                    if len(entry["variants"]) > 1 else "unchanged")
            print(f"    {name:<14} {note:<14} {len(entry['cells']):>3} busy cells"
                  f"   <- {', '.join(entry['variants'])}")

    if multi_cells:
        print(f"\n  Note: {multi_cells} cell(s) in {rooms_multi} room(s) held different"
              f" classes across their sub-rooms.\n"
              f"  Every label was kept and joined with '{args.sep}'."
              f" See the 'Multi-class Cells' sheet.")

    print(f"\n  Written: {out}\n")


if __name__ == "__main__":
    main()
